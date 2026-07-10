## build_outputs.py

"""
build-outputs (Python). Pure rendering — never mutates judgment fields. Writes the three
Excel files FIRST (Apply.xlsx, Review.xlsx, Skip.xlsx), THEN shells out to the unchanged
render_resume.py for Apply-only PDFs. Columns/sort/formatting match build-outputs/SKILL.md.
"""
from __future__ import annotations
import json
import subprocess
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

import config

APPLY_REVIEW_COLS = [
    "jobId", "companyName", "jobTitle", "status", "scorePre", "scorePost",
    "atsScore", "recruiterScore", "keywordsMatched", "keywordsMissing",
    "missingSkills", "skillsRequired", "comment", "outreach", "jobLink",
    "Current Bullets", "Suggested Replacement Bullets",
]
SKIP_COLS = [
    "jobId", "companyName", "jobTitle", "scorePre", "dropReason",
    "missingSkills", "skillsRequired", "comment", "jobLink",
]

HDR_FILL = PatternFill("solid", fgColor="D9D9D9")
APPLY_FILL = PatternFill("solid", fgColor="C6EFCE")   # green
REVIEW_FILL = PatternFill("solid", fgColor="FFEB9C")  # amber
WRAP_TOP = Alignment(wrap_text=True, vertical="top")


def latest_run() -> Path:
    runs = sorted(config.OUTPUTS.glob("run_*"), key=lambda p: p.name)
    if not runs:
        raise FileNotFoundError("No run_* folder found in outputs/.")
    return runs[-1]


def _join_list(v) -> str:
    return "; ".join(v) if isinstance(v, list) else (v or "")


def _bullet_lines(tailored: dict) -> tuple[str, str]:
    """Flatten nested experience groups + projects into row-aligned Current/Suggested text."""
    if not tailored:
        return "", ""
    originals, replacements = [], []
    for role in tailored.get("experience", []):
        for group in role.get("groups", []):
            for b in group.get("bullets", []):
                originals.append(b.get("original", ""))
                rep = b.get("replacement", "")
                replacements.append(rep if b.get("changed") else f"(unchanged) {rep}")
    for p in tailored.get("projects", []):
        originals.append(p.get("original", ""))
        rep = p.get("replacement", "")
        replacements.append(rep if p.get("changed") else f"(unchanged) {rep}")
    return "\n".join(originals), "\n".join(replacements)


def _autofit(ws, wide_cols: set[int]):
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        if col_idx in wide_cols:
            ws.column_dimensions[letter].width = 55
        else:
            longest = max((len(str(c.value or "")) for c in ws[letter]), default=10)
            ws.column_dimensions[letter].width = min(max(longest + 2, 10), 40)


def _write_sheet(path: Path, cols: list[str], rows: list[dict], status_fill: bool):
    wb = Workbook()
    ws = wb.active
    ws.append(cols)
    for c in ws[1]:
        c.font = Font(bold=True)
        c.fill = HDR_FILL
        c.alignment = WRAP_TOP
    ws.freeze_panes = "A2"

    wide = {cols.index(x) + 1 for x in ("comment", "outreach", "keywordsMatched",
            "keywordsMissing", "missingSkills", "skillsRequired",
            "Current Bullets", "Suggested Replacement Bullets") if x in cols}

    for rec in rows:
        cur, sug = _bullet_lines(rec.get("tailored"))
        values = []
        for col in cols:
            if col == "Current Bullets":
                values.append(cur)
            elif col == "Suggested Replacement Bullets":
                values.append(sug)
            elif col == "dropReason":
                values.append((rec.get("stage1") or {}).get("dropReason", "") or "")
            elif col in ("keywordsMatched", "keywordsMissing", "missingSkills"):
                values.append(_join_list(rec.get(col)))
            else:
                values.append(rec.get(col, ""))
        ws.append(values)
        row_i = ws.max_row
        for c in ws[row_i]:
            c.alignment = WRAP_TOP
        if status_fill and "status" in cols:
            st = rec.get("status", "")
            fill = APPLY_FILL if st == "Apply" else REVIEW_FILL if st == "Review" else None
            if fill:
                ws.cell(row=row_i, column=cols.index("status") + 1).fill = fill
        if "jobLink" in cols:
            link = rec.get("jobLink", "")
            if link:
                cell = ws.cell(row=row_i, column=cols.index("jobLink") + 1)
                cell.hyperlink = link
                cell.font = Font(color="0563C1", underline="single")

    _autofit(ws, wide)
    wb.save(str(path))


def run(run_folder: Path | None = None) -> Path:
    folder = run_folder or latest_run()
    records = [json.loads(l) for l in (folder / "evaluations.jsonl")
               .read_text(encoding="utf-8").splitlines() if l.strip()]

    apply = sorted([r for r in records if r.get("status") == "Apply"],
                   key=lambda r: r.get("scorePost", 0), reverse=True)
    review = sorted([r for r in records if r.get("status") == "Review"],
                    key=lambda r: r.get("scorePost", 0), reverse=True)
    skip = [r for r in records if r.get("status") == "Skip"]

    # Excel FIRST, all three, before any PDF.
    _write_sheet(folder / "Apply.xlsx", APPLY_REVIEW_COLS, apply, status_fill=True)
    _write_sheet(folder / "Review.xlsx", APPLY_REVIEW_COLS, review, status_fill=True)
    _write_sheet(folder / "Skip.xlsx", SKIP_COLS, skip, status_fill=False)
    print(f"[build] xlsx: Apply={len(apply)} Review={len(review)} Skip={len(skip)} "
          f"(total {len(records)})")

    # PDFs SECOND — Apply-only, via the unchanged renderer.
    import os
    env = dict(os.environ, RESUME_FONTDIR=config.RESUME_FONTDIR)
    result = subprocess.run(
        [sys.executable, str(config.RENDER_SCRIPT), str(folder)],
        capture_output=True, text=True, env=env,
    )
    print(result.stdout.strip())
    if result.returncode != 0:
        print("[build] renderer STDERR:\n" + result.stderr)
        raise RuntimeError("render_resume.py failed — see stderr above.")
    print(f"[build] done -> {folder}")
    return folder


if __name__ == "__main__":
    run()
