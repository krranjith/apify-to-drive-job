#drive_pipeline.py

"""
Orchestrates the job-match pipeline against Google Drive instead of local CSVs. Two entry
points, called from app.py's Streamlit "Evaluation Pipeline" flow (and usable standalone):

    run_job_pipeline()    — score the latest "Jobs*" file from GOOGLE_DRIVE_ANALYSIS_FOLDER_ID
    run_links_pipeline()  — fetch JDs from the latest "*googlesearch*" file's URLs, write the
                             enriched Jobs-schema data back into that same Drive file, then
                             feed it through the same scoring chain as run_job_pipeline.

Scoring/tailoring/rendering still happen against a local scratch run-folder (config.OUTPUTS) —
only the source data and final deliverables move through Drive.
"""
from __future__ import annotations
import csv
import io
import json
from pathlib import Path
from typing import Callable

import openpyxl

import config
import drive_io
import evaluate
import tailor
import build_outputs
import fetch_links
import jobs_io

_NOOP = lambda msg: None  # noqa: E731

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
PDF_MIME = "application/pdf"


# ---------------------------------------------------------------------------
# Grounding sync
# ---------------------------------------------------------------------------

def sync_grounding_docs(on_progress: Callable[[str], None] = _NOOP) -> None:
    """Download the resume PDF + extended DOCX from GOOGLE_DRIVE_GROUNDING_FOLDER_ID into
    config.RESUME_PDF / config.EXTENDED_DOCX (local cache), if not already present."""
    config.CACHE_DIR.mkdir(parents=True, exist_ok=True)

    resume = drive_io.find_latest(
        config.GROUNDING_FOLDER_ID,
        lambda n: "resume" in n.lower() and n.lower().endswith(".pdf") and "extended" not in n.lower(),
    )
    if resume:
        config.RESUME_PDF.write_bytes(drive_io.download_file(resume["id"], resume["mimeType"]))
        on_progress(f"Synced grounding: {resume['name']} -> {config.RESUME_PDF.name}")
    else:
        on_progress("[warn] no resume PDF found in grounding folder — falling back to canonical JSON")

    extended = drive_io.find_latest(
        config.GROUNDING_FOLDER_ID,
        lambda n: "extended" in n.lower() and n.lower().endswith(".docx"),
    )
    if extended:
        config.EXTENDED_DOCX.write_bytes(drive_io.download_file(extended["id"], extended["mimeType"]))
        on_progress(f"Synced grounding: {extended['name']} -> {config.EXTENDED_DOCX.name}")
    else:
        on_progress("[warn] no extended DOCX found in grounding folder — grounding will be partial")


# ---------------------------------------------------------------------------
# Source parsing
# ---------------------------------------------------------------------------

def build_queue_from_jobs_xlsx(xlsx_bytes: bytes, source_name: str = "jobs_drive") -> list[dict]:
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    header = [str(h or "").strip() for h in next(rows_iter)]
    queue = []
    for values in rows_iter:
        row = {header[i]: ("" if v is None else str(v)) for i, v in enumerate(values) if i < len(header)}
        status = row.get("Status", "")
        if not jobs_io.is_blank_status(status):
            continue
        queue.append(jobs_io.map_to_schema(row, source_name))
    return queue


def build_rows_from_googlesearch_csv(csv_bytes: bytes) -> list[dict]:
    """Parse the Title,Domain,Date,URL,Snippet[,Query] sheet into dicts with a lowercase
    "url" key (fetch_links.fetch_jobs_from_urls reads row["url"])."""
    text = csv_bytes.decode("utf-8", errors="replace")
    reader = csv.DictReader(io.StringIO(text))
    rows = []
    for r in reader:
        row = dict(r)
        row["url"] = row.get("URL", "") or row.get("url", "")
        rows.append(row)
    return rows


# ---------------------------------------------------------------------------
# Shared scoring/tailor/build/upload/writeback chain
# ---------------------------------------------------------------------------

def _run_chain(queue: list[dict], calibrate: int | None, on_progress: Callable[[str], None]) -> dict:
    on_progress(f"Scoring {len(queue)} job(s)…")
    folder = evaluate.run(calibrate=calibrate, queue=queue)

    if calibrate:
        records = [json.loads(l) for l in (folder / "evaluations.jsonl")
                   .read_text(encoding="utf-8").splitlines() if l.strip()]
        counts = {"Apply": 0, "Review": 0, "Skip": 0}
        for r in records:
            counts[r.get("status", "Skip")] = counts.get(r.get("status", "Skip"), 0) + 1
        on_progress(f"Calibration done ({folder.name}) — Apply={counts['Apply']} "
                    f"Review={counts['Review']} Skip={counts['Skip']}. No upload/write-back.")
        return {"run_folder": str(folder), "calibrated": True, "counts": counts}

    on_progress("Tailoring Apply/Review resumes…")
    tailor.run(folder)

    on_progress("Building Excel + PDF outputs…")
    build_outputs.run(folder)

    on_progress(f"Uploading results to Drive folder '{folder.name}'…")
    drive_folder_id = drive_io.create_folder(config.EVALUATED_RESULTS_FOLDER_ID, folder.name)
    for name in ("Apply.xlsx", "Review.xlsx", "Skip.xlsx"):
        path = folder / name
        if path.exists():
            drive_io.upload_file(path, drive_folder_id, mime_type=XLSX_MIME)
    for pdf in sorted(folder.glob("*.pdf")):
        drive_io.upload_file(pdf, drive_folder_id, mime_type=PDF_MIME)

    records = [json.loads(l) for l in (folder / "evaluations.jsonl")
               .read_text(encoding="utf-8").splitlines() if l.strip()]
    counts = {"Apply": 0, "Review": 0, "Skip": 0}
    for r in records:
        counts[r.get("status", "Skip")] = counts.get(r.get("status", "Skip"), 0) + 1

    folder_link = f"https://drive.google.com/drive/folders/{drive_folder_id}"
    on_progress(f"Done — {folder_link}")
    return {
        "run_folder": str(folder),
        "calibrated": False,
        "counts": counts,
        "drive_folder_id": drive_folder_id,
        "drive_folder_link": folder_link,
    }


def _writeback_processed_drive(file_id: str, mime_type: str, scored_urls: set[str],
                                on_progress: Callable[[str], None]) -> None:
    """Set Status=Processed (keyed on applyUrl) on the source Drive file, then re-upload it
    in place so a later run doesn't re-score the same rows."""
    content = drive_io.download_file(file_id, mime_type)
    if mime_type == XLSX_MIME:
        wb = openpyxl.load_workbook(io.BytesIO(content))
        ws = wb.active
        header = [c.value for c in ws[1]]
        try:
            apply_col = header.index("applyUrl") + 1
            status_col = header.index("Status") + 1
        except ValueError:
            on_progress("[warn] source file missing applyUrl/Status column — skipping write-back")
            return
        flipped = 0
        for row in ws.iter_rows(min_row=2):
            url_cell = row[apply_col - 1]
            status_cell = row[status_col - 1]
            if (url_cell.value or "").strip() in scored_urls and not (status_cell.value or "").strip():
                status_cell.value = "Processed"
                flipped += 1
        buf = io.BytesIO()
        wb.save(buf)
        drive_io.update_media(file_id, buf.getvalue(), mime_type)
        on_progress(f"Write-back: flipped {flipped} row(s) to Processed on source Drive file")
    else:
        text = content.decode("utf-8", errors="replace")
        rows = list(csv.reader(io.StringIO(text)))
        if not rows:
            return
        header = rows[0]
        try:
            apply_i = header.index("applyUrl")
            status_i = header.index("Status")
        except ValueError:
            on_progress("[warn] source file missing applyUrl/Status column — skipping write-back")
            return
        flipped = 0
        for row in rows[1:]:
            while len(row) <= max(apply_i, status_i):
                row.append("")
            if row[apply_i].strip() in scored_urls and not row[status_i].strip():
                row[status_i] = "Processed"
                flipped += 1
        out = io.StringIO()
        csv.writer(out).writerows(rows)
        drive_io.update_media(file_id, out.getvalue().encode("utf-8"), "text/csv")
        on_progress(f"Write-back: flipped {flipped} row(s) to Processed on source Drive file")


# ---------------------------------------------------------------------------
# Entry points
# ---------------------------------------------------------------------------

def run_job_pipeline(calibrate: int | None = None,
                      on_progress: Callable[[str], None] = _NOOP) -> dict:
    sync_grounding_docs(on_progress)

    src = drive_io.find_latest(config.ANALYSIS_FOLDER_ID, lambda n: n.lower().startswith("jobs"))
    if not src:
        raise FileNotFoundError("No file starting with 'jobs' found in the analysis Drive folder.")
    on_progress(f"Latest Jobs file: {src['name']}")

    xlsx_bytes = drive_io.download_file(src["id"], src["mimeType"])
    queue = build_queue_from_jobs_xlsx(xlsx_bytes, source_name=src["name"])
    on_progress(f"{len(queue)} row(s) to score (blank Status).")

    result = _run_chain(queue, calibrate, on_progress)

    if not result["calibrated"] and queue:
        scored_urls = {q["applyUrl"] for q in queue if q.get("applyUrl")}
        _writeback_processed_drive(src["id"], src["mimeType"], scored_urls, on_progress)

    return result


def run_links_pipeline(calibrate: int | None = None,
                        on_progress: Callable[[str], None] = _NOOP) -> dict:
    sync_grounding_docs(on_progress)

    src = drive_io.find_latest(config.ANALYSIS_FOLDER_ID, lambda n: "googlesearch" in n.lower())
    if not src:
        raise FileNotFoundError("No file containing 'googlesearch' found in the analysis Drive folder.")
    on_progress(f"Latest googlesearch file: {src['name']}")

    csv_bytes = drive_io.download_file(src["id"], src["mimeType"], export_as="text/csv")
    search_rows = build_rows_from_googlesearch_csv(csv_bytes)
    on_progress(f"Fetching job descriptions for {len(search_rows)} URL(s)…")

    enriched = fetch_links.fetch_jobs_from_urls(search_rows)
    on_progress(f"Fetched: {sum(1 for r in enriched if r['Status'] != 'FETCH_FAILED')} OK, "
                f"{sum(1 for r in enriched if r['Status'] == 'FETCH_FAILED')} failed.")

    # Write the enriched Jobs-schema rows back into the SAME Drive file.
    out = io.StringIO()
    writer = csv.DictWriter(out, fieldnames=config.CSV_COLUMNS)
    writer.writeheader()
    writer.writerows({k: r.get(k, "") for k in config.CSV_COLUMNS} for r in enriched)
    drive_io.update_media(src["id"], out.getvalue().encode("utf-8"), "text/csv")
    on_progress(f"Wrote enriched job data back into '{src['name']}'.")

    queue = [
        jobs_io.map_to_schema(r, src["name"])
        for r in enriched
        if jobs_io.is_blank_status(r.get("Status", ""))
    ]
    on_progress(f"{len(queue)} row(s) to score (fetch succeeded).")

    result = _run_chain(queue, calibrate, on_progress)

    if not result["calibrated"] and queue:
        scored_urls = {q["applyUrl"] for q in queue if q.get("applyUrl")}
        # src stays a native Google Sheet after the CSV re-upload above (Drive convert-imports
        # in place), so write-back must export/re-import through the same original mimeType.
        _writeback_processed_drive(src["id"], src["mimeType"], scored_urls, on_progress)

    return result
